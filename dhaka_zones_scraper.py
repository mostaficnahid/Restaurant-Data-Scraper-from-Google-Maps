"""
============================================================
  Google Maps Restaurant Scraper — Dhaka Premium Zones
  ─────────────────────────────────────────────────────────
  Automatically scrapes restaurant listings from Google Maps
  for three hardcoded Dhaka zones:
      • Banani
      • Baridhara
      • Gulshan

  Results from all three zones are merged into a single,
  colour-coded Excel workbook with one sheet per zone and
  a combined "All Restaurants" sheet.

  HOW TO USE
  ──────────
  1. Install dependencies (one-time):
       pip install selenium webdriver-manager openpyxl

  2. Make sure Google Chrome is installed.

  3. Run:
       python dhaka_zones_scraper.py

  OUTPUT
  ──────
  dhaka_restaurants.xlsx  — 4 sheets:
    • All Restaurants  (every record)
    • Banani
    • Baridhara
    • Gulshan
============================================================
"""

import time
import re
import logging
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
)
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── Logging ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ╔══════════════════════════════════════════════════════╗
# ║               HARDCODED ZONE CONFIG                  ║
# ╚══════════════════════════════════════════════════════╝

# These three zones are fixed — the script will run a
# separate Google Maps search for each zone automatically.
ZONES = [
    "restaurants in Banani, Dhaka",
    "restaurants in Baridhara, Dhaka",
    "restaurants in Gulshan, Dhaka",
]

# Zone short labels (used as Excel sheet names & colour keys)
ZONE_LABELS = ["Banani", "Baridhara", "Gulshan"]

# Unique accent colour per zone (hex, no #)
ZONE_COLORS = {
    "Banani":    "1565C0",   # deep blue
    "Baridhara": "2E7D32",   # deep green
    "Gulshan":   "6A1B9A",   # deep purple
}

CONFIG = {
    "max_results_per_zone": 40,        # restaurants to collect per zone
    "output_file":          "dhaka_restaurants.xlsx",
    "scroll_pause":         2,         # seconds between scroll attempts
    "detail_load_wait":     3,         # seconds to wait on each detail page
    "headless":             True,      # False = visible browser (debug mode)
}


# ╔══════════════════════════════════════════════════════╗
# ║                   DATA MODEL                         ║
# ╚══════════════════════════════════════════════════════╝
@dataclass
class Restaurant:
    zone:          str = "N/A"
    name:          str = "N/A"
    address:       str = "N/A"
    phone:         str = "N/A"
    website:       str = "N/A"
    email:         str = "N/A"
    rating:        str = "N/A"
    review_count:  str = "N/A"
    category:      str = "N/A"
    cuisine:       str = "N/A"
    price_level:   str = "N/A"
    hours:         str = "N/A"
    plus_code:     str = "N/A"
    maps_url:      str = "N/A"
    scraped_at:    str = field(
        default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M")
    )


# Column definitions: (Header label, dataclass field name)
COLUMNS = [
    ("Zone",            "zone"),
    ("Name",            "name"),
    ("Address",         "address"),
    ("Phone",           "phone"),
    ("Website",         "website"),
    ("Email",           "email"),
    ("Rating",          "rating"),
    ("Reviews",         "review_count"),
    ("Category",        "category"),
    ("Cuisine",         "cuisine"),
    ("Price Level",     "price_level"),
    ("Hours",           "hours"),
    ("Plus Code",       "plus_code"),
    ("Google Maps URL", "maps_url"),
    ("Scraped At",      "scraped_at"),
]

# Column widths (index matches COLUMNS order, 1-based)
COL_WIDTHS = [12, 30, 40, 18, 40, 30, 8, 10, 20, 25, 12, 45, 20, 50, 18]


# ╔══════════════════════════════════════════════════════╗
# ║                 BROWSER SETUP                        ║
# ╚══════════════════════════════════════════════════════╝
def build_driver(headless: bool = True) -> webdriver.Chrome:
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--lang=en-US")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)

    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=opts)
    driver.implicitly_wait(5)
    return driver


# ╔══════════════════════════════════════════════════════╗
# ║             SEARCH + SCROLL HELPERS                  ║
# ╚══════════════════════════════════════════════════════╝
def search_google_maps(driver: webdriver.Chrome, query: str) -> None:
    """Navigate to Google Maps and execute the search query."""
    log.info(f"Searching: '{query}'")
    driver.get("https://www.google.com/maps")
    wait = WebDriverWait(driver, 15)
    box  = wait.until(EC.presence_of_element_located((By.ID, "searchboxinput")))
    box.clear()
    box.send_keys(query)
    box.send_keys(Keys.ENTER)
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='feed']")))
    log.info("Results panel loaded.")


def scroll_and_collect_urls(
    driver: webdriver.Chrome,
    max_results: int,
    scroll_pause: float,
) -> list[str]:
    """
    Scroll the results panel and return a deduplicated list
    of Google Maps place URLs (up to max_results).
    """
    FEED_SELECTOR = "div[role='feed']"
    CARD_SELECTOR = "a[href*='/maps/place/']"

    feed = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, FEED_SELECTOR))
    )

    seen: set  = set()
    urls: list = []

    for attempt in range(100):
        for card in driver.find_elements(By.CSS_SELECTOR, CARD_SELECTOR):
            try:
                href = card.get_attribute("href") or ""
                if href and href not in seen:
                    seen.add(href)
                    urls.append(href)
            except StaleElementReferenceException:
                continue

        log.info(f"    Scroll {attempt + 1}: {len(urls)} URLs collected…")

        if len(urls) >= max_results:
            log.info(f"    Target of {max_results} reached.")
            break

        # End-of-results marker
        try:
            driver.find_element(
                By.XPATH, "//*[contains(text(), \"You've reached the end\")]"
            )
            log.info("    End of results reached.")
            break
        except NoSuchElementException:
            pass

        driver.execute_script(
            "arguments[0].scrollTop = arguments[0].scrollHeight", feed
        )
        time.sleep(scroll_pause)

        try:
            feed = driver.find_element(By.CSS_SELECTOR, FEED_SELECTOR)
        except NoSuchElementException:
            log.warning("    Feed panel disappeared.")
            break

    return urls[:max_results]


# ╔══════════════════════════════════════════════════════╗
# ║              DETAIL PAGE PARSER                      ║
# ╚══════════════════════════════════════════════════════╝
def _text(driver, selector, by=By.CSS_SELECTOR) -> str:
    try:
        return driver.find_element(by, selector).text.strip() or "N/A"
    except NoSuchElementException:
        return "N/A"


def parse_detail_page(
    driver: webdriver.Chrome, url: str, zone_label: str
) -> Restaurant:
    """
    Open a Maps place URL and extract all available fields
    into a Restaurant dataclass.
    """
    r = Restaurant(zone=zone_label, maps_url=url)

    try:
        driver.get(url)
        time.sleep(CONFIG["detail_load_wait"])
        wait = WebDriverWait(driver, 10)

        # Name
        try:
            r.name = wait.until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     "h1.DUwDvf, h1[class*='fontHeadlineLarge']")
                )
            ).text.strip()
        except TimeoutException:
            r.name = _text(driver, "h1")

        # Rating
        try:
            r.rating = driver.find_element(
                By.CSS_SELECTOR, "div.F7nice span[aria-hidden='true']"
            ).text.strip()
        except NoSuchElementException:
            pass

        # Review count
        try:
            aria = (
                driver.find_element(
                    By.CSS_SELECTOR,
                    "div.F7nice span[aria-label*='review']"
                ).get_attribute("aria-label") or ""
            )
            nums = re.findall(r"[\d,]+", aria)
            r.review_count = nums[0].replace(",", "") if nums else "N/A"
        except NoSuchElementException:
            pass

        # Category / cuisine
        try:
            type_els = driver.find_elements(
                By.CSS_SELECTOR,
                "button[jsaction*='category'], span.mgr77e"
            )
            types = [el.text.strip() for el in type_els if el.text.strip()]
            if types:
                r.category = types[0]
                r.cuisine  = ", ".join(types)
        except Exception:
            pass

        # Address
        try:
            r.address = driver.find_element(
                By.CSS_SELECTOR,
                "button[data-item-id='address'] .Io6YTe"
            ).text.strip()
        except NoSuchElementException:
            pass

        # Phone
        try:
            r.phone = driver.find_element(
                By.CSS_SELECTOR,
                "button[data-item-id*='phone'] .Io6YTe"
            ).text.strip()
        except NoSuchElementException:
            pass

        # Website
        try:
            r.website = (
                driver.find_element(
                    By.CSS_SELECTOR, "a[data-item-id='authority']"
                ).get_attribute("href") or "N/A"
            )
        except NoSuchElementException:
            pass

        # Hours
        try:
            r.hours = (
                driver.find_element(
                    By.CSS_SELECTOR, "div[aria-label*='Hours'] table, div.t39EBf"
                ).text.strip().replace("\n", " | ")
            )
        except NoSuchElementException:
            try:
                r.hours = driver.find_element(
                    By.CSS_SELECTOR, "button[data-item-id='oh'] .Io6YTe"
                ).text.strip()
            except NoSuchElementException:
                pass

        # Price level
        try:
            txt = driver.find_element(
                By.CSS_SELECTOR,
                "span.mgr77e, span[aria-label*='Price']"
            ).text.strip()
            if re.match(r"^\$+$", txt):
                r.price_level = txt
        except NoSuchElementException:
            pass

        # Plus Code
        try:
            r.plus_code = driver.find_element(
                By.CSS_SELECTOR, "button[data-item-id='oloc'] .Io6YTe"
            ).text.strip()
        except NoSuchElementException:
            pass

        # Email (mailto: links only — rare on Maps)
        try:
            emails = [
                a.get_attribute("href").replace("mailto:", "").strip()
                for a in driver.find_elements(
                    By.CSS_SELECTOR, "a[href^='mailto:']"
                )
                if a.get_attribute("href")
            ]
            if emails:
                r.email = emails[0]
        except Exception:
            pass

    except Exception as exc:
        log.error(f"Error parsing {url}: {exc}")

    return r


# ╔══════════════════════════════════════════════════════╗
# ║                 EXCEL EXPORT                         ║
# ╚══════════════════════════════════════════════════════╝
HEADER_FG  = "FFFFFF"
ALT_BG     = "F3F6FD"
BORDER_CLR = "BDBDBD"


def _border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_sheet(ws, rows: list[Restaurant], accent: str, title: str) -> None:
    """Write a single worksheet with the given accent colour."""
    num_cols = len(COLUMNS)

    # ── Title row ─────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    tc = ws["A1"]
    tc.value     = title
    tc.font      = Font(name="Arial", bold=True, size=13, color=HEADER_FG)
    tc.fill      = PatternFill("solid", fgColor="0D47A1")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header row ────────────────────────────────────────
    h_fill  = PatternFill("solid", fgColor=accent)
    h_font  = Font(name="Arial", bold=True, size=10, color=HEADER_FG)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, (label, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font      = h_font
        c.fill      = h_fill
        c.alignment = h_align
        c.border    = _border()
    ws.row_dimensions[2].height = 22

    # ── Data rows ─────────────────────────────────────────
    d_font  = Font(name="Arial", size=10)
    alt     = PatternFill("solid", fgColor=ALT_BG)
    d_align = Alignment(vertical="top", wrap_text=True)

    for ri, r in enumerate(rows, 3):
        d = asdict(r)
        fill = alt if ri % 2 == 0 else None
        for ci, (_, fname) in enumerate(COLUMNS, 1):
            c = ws.cell(row=ri, column=ci, value=d.get(fname, "N/A"))
            c.font      = d_font
            c.alignment = d_align
            c.border    = _border()
            if fill:
                c.fill = fill
        ws.row_dimensions[ri].height = 40

    # ── Column widths ─────────────────────────────────────
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Freeze & filter ───────────────────────────────────
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}2"


def export_to_excel(
    by_zone: dict[str, list[Restaurant]], filepath: str
) -> None:
    """
    Build a multi-sheet Excel workbook:
      Sheet 1 — All Restaurants (combined)
      Sheet 2 — Banani
      Sheet 3 — Baridhara
      Sheet 4 — Gulshan
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    all_restaurants = [r for zone in by_zone.values() for r in zone]

    # ── Sheet: All Restaurants ────────────────────────────
    ws_all = wb.create_sheet("All Restaurants")
    _write_sheet(
        ws_all,
        all_restaurants,
        accent="37474F",
        title=(
            f"All Zones — Banani · Baridhara · Gulshan  "
            f"|  {datetime.now().strftime('%d %b %Y')}"
        ),
    )

    # ── Per-zone sheets ───────────────────────────────────
    for label in ZONE_LABELS:
        rows   = by_zone.get(label, [])
        accent = ZONE_COLORS.get(label, "1A73E8")
        ws     = wb.create_sheet(label)
        _write_sheet(
            ws,
            rows,
            accent=accent,
            title=(
                f"Restaurants in {label}, Dhaka  "
                f"|  {datetime.now().strftime('%d %b %Y')}"
            ),
        )

    # ── Summary sheet ─────────────────────────────────────
    ws_s = wb.create_sheet("Summary")
    ws_s.column_dimensions["A"].width = 28
    ws_s.column_dimensions["B"].width = 18

    # Header
    for ci, label in enumerate(["Metric", "Count"], 1):
        c = ws_s.cell(row=1, column=ci, value=label)
        c.font      = Font(name="Arial", bold=True, color=HEADER_FG)
        c.fill      = PatternFill("solid", fgColor="37474F")
        c.alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Total Restaurants", len(all_restaurants)),
        ("", ""),
    ]
    for label in ZONE_LABELS:
        summary_rows.append((f"  {label}", len(by_zone.get(label, []))))

    summary_rows += [
        ("", ""),
        ("With Phone Number",  sum(1 for r in all_restaurants if r.phone   != "N/A")),
        ("With Website",       sum(1 for r in all_restaurants if r.website != "N/A")),
        ("With Email",         sum(1 for r in all_restaurants if r.email   != "N/A")),
        ("With Rating",        sum(1 for r in all_restaurants if r.rating  != "N/A")),
        ("", ""),
        ("Scraped At", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    for ri, (metric, val) in enumerate(summary_rows, 2):
        ws_s.cell(row=ri, column=1, value=metric).font = Font(
            name="Arial", bold=bool(metric and not metric.startswith("  ")), size=10
        )
        ws_s.cell(row=ri, column=2, value=val).font = Font(name="Arial", size=10)

    wb.save(filepath)
    log.info(f"Workbook saved → {filepath}")


# ╔══════════════════════════════════════════════════════╗
# ║                     MAIN                             ║
# ╚══════════════════════════════════════════════════════╝
def main():
    driver = build_driver(headless=CONFIG["headless"])

    # Collect results keyed by zone label
    by_zone: dict[str, list[Restaurant]] = {label: [] for label in ZONE_LABELS}

    try:
        for query, label in zip(ZONES, ZONE_LABELS):
            log.info(f"\n{'═'*55}")
            log.info(f"  ZONE: {label}")
            log.info(f"{'═'*55}")

            search_google_maps(driver, query)
            urls = scroll_and_collect_urls(
                driver,
                CONFIG["max_results_per_zone"],
                CONFIG["scroll_pause"],
            )
            log.info(f"  {len(urls)} unique URLs found for {label}.")

            for i, url in enumerate(urls, 1):
                log.info(f"  [{i}/{len(urls)}] Scraping…")
                r = parse_detail_page(driver, url, zone_label=label)
                by_zone[label].append(r)
                log.info(f"    ✓ {r.name} | {r.phone}")

    finally:
        driver.quit()

    total = sum(len(v) for v in by_zone.values())
    if total == 0:
        log.warning("No data collected. Check your network or try headless=False.")
        return

    export_to_excel(by_zone, CONFIG["output_file"])

    # ── Final summary ─────────────────────────────────────
    print(f"\n{'─'*55}")
    print(f"  Dhaka Scrape Complete")
    print(f"{'─'*55}")
    for label in ZONE_LABELS:
        print(f"  {label:<12} → {len(by_zone[label])} restaurants")
    print(f"  {'TOTAL':<12} → {total} restaurants")
    print(f"\n  Saved to: {CONFIG['output_file']}")
    print(f"{'─'*55}\n")


if __name__ == "__main__":
    main()
